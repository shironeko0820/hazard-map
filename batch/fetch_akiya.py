"""
空き家率データ取得・集計スクリプト

データソース: 令和5年（2023年）住宅・土地統計調査
  基準日: 2023年10月1日
  公表日: 2024年9月25日
  出典: e-Stat（総務省統計局）

取得ファイル:
  - 表1-2: 住宅の種類別住宅数（statInfId: 000040209842）
      列: 地域分類コード, 地域名, 総住宅数, 居住世帯あり, ...
  - 表34-2: 空き家の種類別空き家数（statInfId: 000040209901）
      列: 地域分類コード, 地域名, 腐朽破損, 建て方, 構造, 総数, ...

対象: 市・区および人口1万5千人以上の町村（地域分類コード='1'または'0'以外）
出力: public/akiya.json  {城市コード5桁: {akiya_rate, total_housing, vacant_housing, name}}
"""

import json
import os
import io
import requests
import openpyxl

BASE_DIR = os.path.dirname(__file__)
OUTPUT_PATH = os.path.join(BASE_DIR, "..", "public", "akiya.json")

DOWNLOAD_URL = "https://www.e-stat.go.jp/stat-search/file-download?statInfId={}&fileKind=0"
TOTAL_HOUSING_ID = "000040209842"   # 表1-2: 住宅の種類別住宅数（市区町村）
VACANT_HOUSING_ID = "000040209901"  # 表34-2: 空き家の種類別空き家数（市区町村）

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; MachiScore/1.0)"}


def download_xlsx(stat_inf_id: str) -> openpyxl.Workbook:
    url = DOWNLOAD_URL.format(stat_inf_id)
    print(f"  ダウンロード: {url}")
    resp = requests.get(url, headers=HEADERS, timeout=120)
    resp.raise_for_status()
    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)


def parse_code(area_str: str) -> str | None:
    """'01100_札幌市' → '01100'（5桁）"""
    if not area_str:
        return None
    s = str(area_str).strip()
    if "_" in s:
        code = s.split("_")[0]
    else:
        code = s
    code = code.strip().replace(",", "")
    if code.isdigit() and len(code) == 5:
        return code
    return None


def extract_total_housing(wb: openpyxl.Workbook) -> dict[str, int]:
    """
    表1-2 から {市区町村コード: 総住宅数} を抽出。

    シート構造（e001_2.xlsx）:
      行1-8: ヘッダー
      行9:   列ラベル行 ('地域分類コード', '地域名', ' ', ...)
      行10-: データ  (col0=地域分類コード, col1=地域名XXXXX_NAME, col2=総住宅数)
    """
    ws = wb.active
    result: dict[str, int] = {}

    # データ開始行を検出（'地域分類コード' が現れた次の行）
    data_start = 10  # デフォルト
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row[0] == "地域分類コード" or (row[0] and "コード" in str(row[0])):
            data_start = i + 1
            break

    print(f"    総住宅数 データ開始行: {data_start}")

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        region_type = str(row[0] or "").strip()
        # 'a' = 全国・都道府県（集計対象外）, '0'/'1' = 市区町村
        if region_type == "a":
            continue
        code = parse_code(str(row[1] or ""))
        if not code or code == "00000":
            continue
        name = str(row[1] or "").split("_", 1)[-1] if "_" in str(row[1] or "") else ""
        total = row[2]
        if total is not None:
            try:
                v = int(float(str(total).replace(",", "")))
                if v > 0:
                    result[code] = v
            except (ValueError, TypeError):
                pass

    print(f"    → {len(result):,}市区町村")
    return result


def extract_vacant_housing(wb: openpyxl.Workbook) -> dict[str, int]:
    """
    表34-2 から {市区町村コード: 空き家数（総数）} を抽出。

    シート構造（e034_2.xlsx）:
      行1-8:  ヘッダー
      行9:    列ラベル行 ('地域分類コード', '地域名', '腐朽破損の有無', '建て方', '構造', 0_総数, ...)
      行10-:  データ  各行は (地域×腐朽破損×建て方×構造) の組み合わせ

    '0_総数' × '0_総数' × '0_総数' の行（全組み合わせの合計）だけ抽出する。
    """
    ws = wb.active
    result: dict[str, int] = {}

    data_start = 10
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if row[0] == "地域分類コード" or (row[0] and "コード" in str(row[0])):
            data_start = i + 1
            break

    print(f"    空き家数 データ開始行: {data_start}")

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        region_type = str(row[0] or "").strip()
        if region_type == "a":
            continue
        code = parse_code(str(row[1] or ""))
        if not code or code == "00000":
            continue

        # col2=腐朽破損, col3=建て方, col4=構造 — すべて '0_総数' の行のみ
        c2 = str(row[2] or "").strip()
        c3 = str(row[3] or "").strip()
        c4 = str(row[4] or "").strip()
        if not (c2.startswith("0") and c3.startswith("0") and c4.startswith("0")):
            continue

        vacant = row[5]  # col5 = 0_総数 (空き家数合計)
        if vacant is not None:
            try:
                v = int(float(str(vacant).replace(",", "")))
                result[code] = v
            except (ValueError, TypeError):
                pass

    print(f"    → {len(result):,}市区町村")
    return result


def main():
    print("=== 空き家率データ取得開始 ===\n")

    print("[1/2] 総住宅数 (表1-2) 取得")
    wb_total = download_xlsx(TOTAL_HOUSING_ID)
    total_map = extract_total_housing(wb_total)

    print("\n[2/2] 空き家数 (表34-2) 取得")
    wb_vacant = download_xlsx(VACANT_HOUSING_ID)
    vacant_map = extract_vacant_housing(wb_vacant)

    print("\n=== 空き家率計算 ===")
    result: dict[str, dict] = {}
    for code, total in total_map.items():
        vacant = vacant_map.get(code)
        if vacant is None or total <= 0:
            continue
        rate = round(vacant / total * 100, 2)
        result[code] = {
            "akiya_rate": rate,
            "total_housing": total,
            "vacant_housing": vacant,
        }

    print(f"集計完了: {len(result):,}市区町村")
    if result:
        rates = [v["akiya_rate"] for v in result.values()]
        print(f"空き家率: min={min(rates):.1f}% / avg={sum(rates)/len(rates):.1f}% / max={max(rates):.1f}%")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n出力: {OUTPUT_PATH}")
    print("=== 完了 ===")


if __name__ == "__main__":
    main()
